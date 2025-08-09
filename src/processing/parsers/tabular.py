# Tabular parsers (CSV/TSV/XLSX)
from pathlib import Path
from typing import Dict, List
import csv

from ..io import ensure_output_dir, write_text_file
from ..normalize import normalize_newlines, join_columns_to_tabs


def _row_to_tabs(row: List[str]) -> str:
    # Delegate to shared normalization helper
    return join_columns_to_tabs(row)


def _write_out(base_dir: Path, stem: str, text: str) -> Path:
    out_dir = ensure_output_dir(base_dir)
    out_path = out_dir / (stem + ".txt")
    write_text_file(out_path, text)
    return out_path


def parse_csv(src_path: Path, base_dir: Path) -> Dict[str, object]:
    p = Path(src_path)
    base = Path(base_dir)
    lines: List[str] = []
    rows_count = 0
    with open(p, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            # Avoid list comprehension, build explicitly
            line = _row_to_tabs(row)
            lines.append(line)
            rows_count = rows_count + 1
    text = "\n".join(lines)
    text = normalize_newlines(text)
    out_path = _write_out(base, p.stem, text)
    return {"out_path": str(out_path), "text": text, "rows": rows_count}


def parse_tsv(src_path: Path, base_dir: Path) -> Dict[str, object]:
    p = Path(src_path)
    base = Path(base_dir)
    lines: List[str] = []
    rows_count = 0
    with open(p, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f, delimiter="\t")
        for row in reader:
            line = _row_to_tabs(row)
            lines.append(line)
            rows_count = rows_count + 1
    text = "\n".join(lines)
    text = normalize_newlines(text)
    out_path = _write_out(base, p.stem, text)
    return {"out_path": str(out_path), "text": text, "rows": rows_count}


def parse_xlsx(src_path: Path, base_dir: Path) -> Dict[str, object]:
    from openpyxl import load_workbook  # local import to keep optional

    p = Path(src_path)
    base = Path(base_dir)
    wb = load_workbook(filename=str(p), read_only=True, data_only=True)

    lines: List[str] = []
    sheets_meta: List[Dict[str, object]] = []

    si = 0
    while si < len(wb.sheetnames):
        name = wb.sheetnames[si]
        ws = wb[name]
        # Sheet separator
        lines.append("=== Sheet: " + name + " ===")
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            # Convert row to list of strings
            cols: List[str] = []
            ci = 0
            # row is a tuple
            while ci < len(row):
                val = row[ci]
                if val is None:
                    cols.append("")
                else:
                    cols.append(str(val))
                ci = ci + 1
            line = _row_to_tabs(cols)
            lines.append(line)
            row_count = row_count + 1
        sheets_meta.append({"name": name, "rows": row_count})
        si = si + 1

    text = "\n".join(lines)
    text = normalize_newlines(text)
    out_path = _write_out(base, p.stem, text)
    return {"out_path": str(out_path), "text": text, "sheets": sheets_meta}
